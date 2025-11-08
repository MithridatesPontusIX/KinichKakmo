"""
KinichKakmo - ICIJ Offshore Leaks Search Application
Sophisticated interface for searching offshore leaks database
"""

import requests
import streamlit as st
import pandas as pd
import sqlite3
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
import os
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib import colors
from typing import List, Dict, Any, Optional, Tuple

# ============================================================================
# How to run it!!!
# ============================================================================


# 1. Create virtual environment
# python3 -m venv .venv
# 2. Activate environment
# macOS/Linux:
# source .venv/bin/activate
# Windows (PowerShell):
# .venv\Scripts\activate
# 3. Install dependencies
# pip install -r requirements.txt


# ============================================================================
# CONFIGURATION
# ============================================================================

DB_FILE = "kinichkakmo.db"
ICIJ_API_URL = "https://offshoreleaks.icij.org/api/v1/reconcile"
ICIJ_NODE_URL = "https://offshoreleaks.icij.org/nodes/"

# Color Palette - Red, White, Gray only
COLORS = {
    'primary_red': '#C62828',
    'dark_red': '#B71C1C',
    'light_red': '#E53935',
    'white': '#FFFFFF',
    'light_gray': '#F5F5F5',
    'medium_gray': '#9E9E9E',
    'dark_gray': '#424242',
    'border_gray': '#E0E0E0'
}

# Available data sources
DATA_SOURCES = [
    "Panama Papers",
    "Paradise Papers",
    "Pandora Papers",
    "Bahamas Leaks",
    "Offshore Leaks"
]

# ============================================================================
# DATABASE FUNCTIONS
# ============================================================================

def init_database() -> None:
    """Initialize SQLite database with required tables."""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    # Search history table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS search_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            query TEXT NOT NULL,
            sources TEXT,
            results_count INTEGER,
            search_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    
    # Saved searches table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS saved_searches (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            query TEXT NOT NULL,
            sources TEXT,
            notes TEXT,
            created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    
    conn.commit()
    conn.close()


def save_search_history(query: str, sources: List[str], results_count: int) -> None:
    """Save search to history."""
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        sources_str = ','.join(sources) if sources else ''
        cursor.execute(
            "INSERT INTO search_history (query, sources, results_count) VALUES (?, ?, ?)",
            (query, sources_str, results_count)
        )
        conn.commit()
        conn.close()
    except Exception as e:
        st.error(f"Error saving search history: {e}")


def get_search_history(limit: int = 10) -> List[Tuple]:
    """Retrieve recent search history."""
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute(
            "SELECT query, sources, results_count, search_date FROM search_history ORDER BY search_date DESC LIMIT ?",
            (limit,)
        )
        history = cursor.fetchall()
        conn.close()
        return history
    except:
        return []


def save_search(name: str, query: str, sources: List[str], notes: str = "") -> bool:
    """Save a search for later use."""
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        sources_str = ','.join(sources) if sources else ''
        cursor.execute(
            "INSERT INTO saved_searches (name, query, sources, notes) VALUES (?, ?, ?, ?)",
            (name, query, sources_str, notes)
        )
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        st.error(f"Error saving search: {e}")
        return False


def get_saved_searches() -> List[Tuple]:
    """Retrieve all saved searches."""
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id, name, query, sources, notes, created_date FROM saved_searches ORDER BY created_date DESC"
        )
        saved = cursor.fetchall()
        conn.close()
        return saved
    except:
        return []


def delete_saved_search(search_id: int) -> bool:
    """Delete a saved search."""
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute("DELETE FROM saved_searches WHERE id = ?", (search_id,))
        conn.commit()
        conn.close()
        return True
    except:
        return False


# ============================================================================
# EXPORT FUNCTIONS
# ============================================================================

def create_excel_export(df_export: pd.DataFrame, query: str, sources: List[str]) -> bytes:
    """Create formatted Excel export."""
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Search Results"
    
    header_fill = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    ws['A1'] = "ICIJ Offshore Leaks Search Results"
    ws['A1'].font = Font(size=16, bold=True, color="C62828")
    ws['A2'] = f"Query: {query}"
    ws['A3'] = f"Sources: {', '.join(sources) if sources else 'All'}"
    ws['A4'] = f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    
    start_row = 6
    for col_num, column_title in enumerate(df_export.columns, 1):
        cell = ws.cell(row=start_row, column=col_num)
        cell.value = column_title
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for row_num, row_data in enumerate(df_export.values, start_row + 1):
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=cell_value)
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def create_pdf_export(df_export: pd.DataFrame, query: str, sources: List[str]) -> bytes:
    """Create professional PDF report."""
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=letter)
    story = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#C62828'),
        spaceAfter=30,
        fontName='Helvetica-Bold'
    )
    
    story.append(Paragraph("ICIJ Offshore Leaks Search Report", title_style))
    story.append(Spacer(1, 0.2*inch))
    
    info_style = styles['Normal']
    story.append(Paragraph(f"<b>Search Query:</b> {query}", info_style))
    story.append(Paragraph(f"<b>Sources:</b> {', '.join(sources) if sources else 'All'}", info_style))
    story.append(Paragraph(f"<b>Results Found:</b> {len(df_export)}", info_style))
    story.append(Paragraph(f"<b>Report Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", info_style))
    story.append(Spacer(1, 0.3*inch))
    
    table_data = [df_export.columns.tolist()] + df_export.values.tolist()
    
    col_widths = [2*inch, 0.8*inch, 1.5*inch]
    if len(df_export.columns) > 3:
        col_widths.extend([1.2*inch] * (len(df_export.columns) - 3))
    
    table = Table(table_data, colWidths=col_widths[:len(df_export.columns)])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#C62828')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    
    story.append(table)
    doc.build(story)
    output.seek(0)
    return output.getvalue()


# ============================================================================
# API FUNCTIONS
# ============================================================================

def search_icij_database(query: str, timeout: int = 15) -> Optional[List[Dict]]:
    """Search the ICIJ database and return results."""
    try:
        payload = {
            "type": "Entity",
            "queries": {
                "q0": {
                    "query": query
                }
            }
        }
        response = requests.post(ICIJ_API_URL, json=payload, timeout=timeout)
        response.raise_for_status()
        data = response.json()
        
        if "q0" in data and "result" in data["q0"] and data["q0"]["result"]:
            return data["q0"]["result"]
        return None
    except requests.exceptions.Timeout:
        st.error("Search timed out. Please try again.")
        return None
    except requests.exceptions.RequestException:
        st.error("Error connecting to ICIJ database. Please check your internet connection.")
        return None
    except Exception as e:
        st.error(f"An unexpected error occurred: {str(e)}")
        return None


# ============================================================================
# FILTER FUNCTIONS
# ============================================================================

def apply_filters(
    results: List[Dict],
    sources: List[str],
    entity_type: str,
    min_score: int,
    jurisdiction: str,
    date_range: str
) -> List[Dict]:
    """Apply various filters to search results."""
    filtered = results
    
    # Source filter
    if sources:
        filtered = [
            r for r in filtered 
            if any(source.lower() in r.get('description', '').lower() 
                  or source.lower() in r.get('name', '').lower() 
                  for source in sources)
        ]
    
    # Entity type filter
    if entity_type != "All":
        filtered = [
            r for r in filtered
            if r.get('types') and len(r.get('types', [])) > 0 
            and r.get('types')[0].get('name', '') == entity_type
        ]
    
    # Score filter
    if min_score > 0:
        filtered = [r for r in filtered if r.get('score', 0) >= min_score]
    
    # Jurisdiction filter
    if jurisdiction:
        filtered = [
            r for r in filtered
            if jurisdiction.lower() in r.get('description', '').lower() or
               jurisdiction.lower() in r.get('name', '').lower()
        ]
    
    # Date range filter
    if date_range != "All Time":
        date_source_map = {
            "2021-Present (Pandora)": ["Pandora Papers"],
            "2016-2017 (Panama/Paradise/Bahamas)": ["Panama Papers", "Paradise Papers", "Bahamas Leaks"],
            "2013 (Offshore Leaks)": ["Offshore Leaks"]
        }
        allowed_sources = date_source_map.get(date_range, [])
        if allowed_sources:
            filtered = [
                r for r in filtered
                if any(src.lower() in r.get('description', '').lower() for src in allowed_sources)
            ]
    
    return filtered


# ============================================================================
# UI STYLING
# ============================================================================

def load_custom_css():
    """Load custom CSS for Gothic styling with Red, White, Gray theme."""
    st.markdown(
        f"""
        <style>
        @import url('https://fonts.googleapis.com/css2?family=Cinzel:wght@400;600;700&family=Gothic+A1:wght@300;400;600;700&display=swap');
        @import url('https://fonts.googleapis.com/icon?family=Material+Icons');
        
        .material-icons {{
            font-family: 'Material Icons';
            font-weight: normal;
            font-style: normal;
            font-size: 24px;
            display: inline-block;
            line-height: 1;
            text-transform: none;
            letter-spacing: normal;
            word-wrap: normal;
            white-space: nowrap;
            direction: ltr;
            vertical-align: middle;
        }}
        
        html, body, [class*="css"] {{
            font-family: 'Gothic A1', sans-serif;
        }}
        
        .stApp {{
            background-color: {COLORS['dark_gray']};
        }}
        
        h1, h2, h3, h4, h5, h6 {{
            font-family: 'Cinzel', serif !important;
            color: {COLORS['primary_red']} !important;
        }}
        
        .main-header {{
            background: linear-gradient(135deg, {COLORS['primary_red']} 0%, {COLORS['dark_red']} 100%);
            padding: 3rem 2rem;
            border-radius: 0;
            margin: -1rem -1rem 2rem -1rem;
            box-shadow: 0 6px 12px rgba(0,0,0,0.15);
            text-align: center;
        }}
        
        .main-header h1 {{
            color: {COLORS['dark_gray']} !important;
            margin: 0;
            font-size: 3.5rem;
            font-weight: 700;
            letter-spacing: 2px;
            text-transform: uppercase;
        }}
        
        .main-header p {{
            color: {COLORS['dark_gray']} !important;
            margin: 1rem 0 0 0;
            font-size: 1.2rem;
            letter-spacing: 1px;
            font-weight: 300;
        }}
        
        .result-card {{
            background-color: {COLORS['dark_gray']};
            border-left: 6px solid {COLORS['primary_red']};
            padding: 2rem;
            margin: 1.5rem 0;
            border-radius: 0;
            box-shadow: 0 3px 6px rgba(0,0,0,0.1);
            transition: all 0.3s ease;
        }}
        
        .result-card:hover {{
            transform: translateX(5px);
            box-shadow: 0 6px 12px rgba(0,0,0,0.15);
            border-left-width: 8px;
        }}
        
        .result-title {{
            color: {COLORS['dark_red']};
            font-size: 1.5rem;
            font-weight: 700;
            margin-bottom: 1rem;
            font-family: 'Cinzel', serif;
            letter-spacing: 1px;
        }}
        
        .result-meta {{
            color: {COLORS['light_gray']};
            font-size: 0.95rem;
            margin: 0.5rem 0;
            font-weight: 400;
        }}
        
        .source-badge {{
            display: inline-block;
            background-color: {COLORS['dark_gray']};
            color: {COLORS['dark_gray']};
            padding: 0.3rem 0.8rem;
            border-radius: 0;
            font-size: 0.8rem;
            margin: 0.3rem 0.3rem 0.3rem 0;
            font-weight: 600;
            letter-spacing: 0.5px;
            text-transform: uppercase;
        }}
        
        .view-link {{
            display: inline-block;
            background-color: {COLORS['primary_red']};
            color: {COLORS['dark_gray']} !important;
            padding: 0.8rem 1.5rem;
            border-radius: 0;
            text-decoration: none;
            margin-top: 1rem;
            font-weight: 600;
            letter-spacing: 1px;
            text-transform: uppercase;
            transition: all 0.3s ease;
            border: 2px solid {COLORS['primary_red']};
        }}
        
        .view-link:hover {{
            background-color: {COLORS['dark_red']};
            border-color: {COLORS['dark_red']};
            text-decoration: none;
            transform: translateY(-2px);
        }}
        
        .stats-box {{
            background-color: {COLORS['light_gray']};
            padding: 2rem 1rem;
            border-radius: 0;
            text-align: center;
            box-shadow: 0 3px 6px rgba(0,0,0,0.1);
            border-top: 4px solid {COLORS['primary_red']};
        }}
        
        .stats-number {{
            font-size: 3rem;
            font-weight: 700;
            color: {COLORS['primary_red']};
            font-family: 'Cinzel', serif;
            letter-spacing: 2px;
        }}
        
        .stats-label {{
            color: {COLORS['medium_gray']};
            font-size: 0.95rem;
            margin-top: 0.5rem;
            text-transform: uppercase;
            letter-spacing: 1px;
            font-weight: 600;
        }}
        
        .stButton > button {{
            background-color: {COLORS['primary_red']};
            color: {COLORS['dark_gray']};
            border: none;
            padding: 0.8rem 2rem;
            border-radius: 0;
            font-weight: 600;
            letter-spacing: 1px;
            text-transform: uppercase;
            transition: all 0.3s ease;
            font-family: 'Gothic A1', sans-serif;
        }}
        
        .stButton > button:hover {{
            background-color: {COLORS['dark_red']};
            transform: translateY(-2px);
            box-shadow: 0 4px 8px rgba(0,0,0,0.2);
        }}
        
        .stTextInput > div > div > input {{
            border-radius: 0;
            border: 2px solid {COLORS['border_gray']};
            padding: 0.8rem;
            font-family: 'Gothic A1', sans-serif;
        }}
        
        .stTextInput > div > div > input:focus {{
            border-color: {COLORS['primary_red']};
            box-shadow: 0 0 0 2px rgba(198, 40, 40, 0.1);
        }}
        
        .comparison-card {{
            background-color: {COLORS['dark_gray']};
            padding: 1.5rem;
            border-radius: 0;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            margin: 1rem 0;
            border-left: 4px solid {COLORS['medium_gray']};
        }}
        
        .sidebar-section {{
            background-color: {COLORS['dark_gray']};
            padding: 1.5rem;
            border-radius: 0;
            margin-bottom: 1.5rem;
            border-left: 4px solid {COLORS['primary_red']};
        }}
        
        .stTabs [data-baseweb="tab-list"] {{
            gap: 0;
        }}
        
        .stTabs [data-baseweb="tab"] {{
            background-color: {COLORS['dark_gray']};
            border-radius: 0;
            color: {COLORS['dark_gray']};
            font-weight: 600;
            padding: 1rem 2rem;
            border: 2px solid {COLORS['border_gray']};
            border-bottom: none;
            font-family: 'Gothic A1', sans-serif;
            text-transform: uppercase;
            letter-spacing: 1px;
        }}
        
        .stTabs [aria-selected="true"] {{
            background-color: {COLORS['primary_red']};
            color: {COLORS['dark_gray']} !important;
            border-color: {COLORS['primary_red']};
        }}
        
        .icon-text {{
            display: inline-flex;
            align-items: center;
            gap: 0.5rem;
        }}
        
        div[data-testid="stExpander"] {{
            background-color: {COLORS['dark_gray']};
            border-radius: 0;
            border: 2px solid {COLORS['border_gray']};
        }}
        
        .stSelectbox > div > div {{
            border-radius: 0;
            border-color: {COLORS['border_gray']};
        }}
        
        .stMultiSelect > div > div {{
            border-radius: 0;
            border-color: {COLORS['border_gray']};
        }}
        </style>
        """,
        unsafe_allow_html=True
    )


# ============================================================================
# MAIN APPLICATION
# ============================================================================

def main():
    """Main application entry point."""
    
    # Page configuration
    st.set_page_config(
        page_title="KinichKakmo - ICIJ Offshore Leaks Search",
        page_icon="M",
        layout="wide",
        initial_sidebar_state="expanded"
    )
    
    # Initialize database
    init_database()
    
    # Load custom styling
    load_custom_css()
    
    # Initialize session state
    if 'comparison_list' not in st.session_state:
        st.session_state.comparison_list = []
    
    # Header
    st.markdown(
        """
        <div class="main-header">
            <h1>KINICHKAKMO</h1>
            <p>ICIJ Offshore Leaks Database Search System</p> 
        </div>
        """,
        unsafe_allow_html=True
    )
    
    # Sidebar
    with st.sidebar:
        st.markdown('<div class="sidebar-section">', unsafe_allow_html=True)
        st.markdown("### About KinichKakmo")
        st.info(
            "Why did the Socialist bring extra snacks to the meeting?"
            " Because in Socialism, everyone gets a bite.. "
        )
        st.markdown('</div>', unsafe_allow_html=True)
        
        with st.expander("How to Use", expanded=False):
            st.markdown("""
            **Search Process:**
            1. Enter search terms (company, person, or location)
            2. Apply source filters as needed
            3. Configure advanced filters for precision
            4. Review comprehensive results
            5. Export data in preferred format
            6. Save searches for future reference
            """)
        
        with st.expander("Example Searches", expanded=False):
            st.markdown("""
            **Corporate Entities:**
            - Mossack Fonseca
            - HSBC Holdings
            - Atlas Network
            
            **Geographic Locations:**
            - British Virgin Islands
            - Panama
            - Cayman Islands
            
            **Investigation Names:**
            - Panama Papers
            - Paradise Papers
            - Pandora Papers
            """)
            
        with st.expander("Data Sources", expanded=False):
            st.markdown("""
            - **Panama Papers** (2016): 11.5M documents
            - **Paradise Papers** (2017): 13.4M documents
            - **Pandora Papers** (2021): 11.9M documents
            - **Bahamas Leaks** (2016): 1.3M documents
            - **Offshore Leaks** (2013): 2.5M documents
            """)
    
    # Main tabs
    tab1, tab2, tab3, tab4 = st.tabs(["SEARCH", "VISUALIZATIONS", "COMPARE", "SAVED & HISTORY"])
    
    # ========================================================================
    # TAB 1: SEARCH
    # ========================================================================
    with tab1:
        col1, col2 = st.columns([2, 1])
        
        with col1:
            query = st.text_input(
                "Search Query",
                placeholder="Enter company name, person, or location...",
                help="Search the ICIJ database for entities, officers, or intermediaries",
                label_visibility="collapsed"
            )
        
        with col2:
            sources = st.multiselect(
                "Filter by Source",
                DATA_SOURCES,
                help="Select specific investigations to filter results",
                label_visibility="collapsed",
                placeholder="All Sources"
            )
        
        # Advanced filters
        with st.expander("Advanced Filters", expanded=False):
            col_f1, col_f2, col_f3 = st.columns(3)
            
            with col_f1:
                entity_type_filter = st.selectbox(
                    "Entity Type",
                    ["All", "Officer", "Entity", "Intermediary", "Address"],
                    help="Filter by entity type"
                )
            
            with col_f2:
                min_score = st.slider(
                    "Minimum Match Score",
                    0, 100, 0,
                    help="Only show results with match score above this value"
                )
            
            with col_f3:
                max_results = st.number_input(
                    "Maximum Results",
                    min_value=5, max_value=100, value=20,
                    help="Limit number of displayed results"
                )
            
            col_f4, col_f5 = st.columns(2)
            
            with col_f4:
                jurisdiction_filter = st.text_input(
                    "Jurisdiction Filter",
                    placeholder="e.g., Panama, British Virgin Islands",
                    help="Filter by jurisdiction or country"
                )
            
            with col_f5:
                date_range_filter = st.selectbox(
                    "Data Source Period",
                    ["All Time", "2021-Present (Pandora)", "2016-2017 (Panama/Paradise/Bahamas)", "2013 (Offshore Leaks)"],
                    help="Filter by data source release year"
                )
        
        # Execute search
        if query:
            with st.spinner(f'Searching for "{query}"...'):
                results = search_icij_database(query)
                
                if results:
                    # Apply filters
                    filtered_results = apply_filters(
                        results,
                        sources,
                        entity_type_filter,
                        min_score,
                        jurisdiction_filter,
                        date_range_filter
                    )
                    
                    filtered_results = filtered_results[:max_results]
                    
                    # Save to history
                    save_search_history(query, sources, len(filtered_results))
                    
                    if filtered_results:
                        st.markdown("---")
                        
                        # Statistics
                        col_stat1, col_stat2, col_stat3 = st.columns(3)
                        with col_stat1:
                            st.markdown(
                                f"""
                                <div class="stats-box">
                                    <div class="stats-number">{len(filtered_results)}</div>
                                    <div class="stats-label">Results Found</div>
                                </div>
                                """,
                                unsafe_allow_html=True
                            )
                        with col_stat2:
                            st.markdown(
                                f"""
                                <div class="stats-box">
                                    <div class="stats-number">{len(sources) if sources else 'All'}</div>
                                    <div class="stats-label">Sources Selected</div>
                                </div>
                                """,
                                unsafe_allow_html=True
                            )
                        with col_stat3:
                            avg_score = sum(r.get('score', 0) for r in filtered_results) / len(filtered_results)
                            st.markdown(
                                f"""
                                <div class="stats-box">
                                    <div class="stats-number">{avg_score:.1f}</div>
                                    <div class="stats-label">Avg Match Score</div>
                                </div>
                                """,
                                unsafe_allow_html=True
                            )
                        
                        st.markdown("### Search Results")
                        
                        # Display results
                        for idx, r in enumerate(filtered_results, 1):
                            entity_name = r.get('name', 'Unknown Entity')
                            entity_id = r.get('id', '')
                            match_score = r.get('score', 0)
                            description = r.get('description', '')
                            entity_type = r.get('types', [{}])[0].get('name', 'Entity') if r.get('types') else 'Entity'
                            
                            match_quality = "HIGH" if match_score >= 80 else "MEDIUM" if match_score >= 50 else "LOW"
                            
                            source_tags = ""
                            if sources:
                                for source in sources:
                                    if source.lower() in description.lower() or source.lower() in entity_name.lower():
                                        source_tags += f'<span class="source-badge">{source}</span>'
                            
                            col_result, col_compare = st.columns([10, 1])
                            
                            with col_result:
                                st.markdown(
                                    f"""
                                    <div class="result-card">
                                        <div class="result-title">{idx}. {entity_name}</div>
                                        <div class="result-meta"><strong>Type:</strong> {entity_type}</div>
                                        <div class="result-meta"><strong>Match Quality:</strong> {match_quality} ({match_score:.1f}/100)</div>
                                        <div class="result-meta"><strong>Entity ID:</strong> {entity_id}</div>
                                        {f'<div class="result-meta"><strong>Description:</strong> {description}</div>' if description else ''}
                                        <div style="margin-top: 1rem;">{source_tags}</div>
                                        <a href="{ICIJ_NODE_URL}{entity_id}" target="_blank" class="view-link">
                                            View Details
                                        </a>
                                    </div>
                                    """,
                                    unsafe_allow_html=True
                                )
                            
                            with col_compare:
                                if st.button("Add", key=f"add_{idx}", help="Add to comparison"):
                                    if r not in st.session_state.comparison_list:
                                        st.session_state.comparison_list.append(r)
                                        st.success("Added")
                        
                        st.markdown("---")
                        
                        # Prepare export data
                        df = pd.DataFrame(filtered_results)
                        df_export = df[['name', 'score', 'id']].copy()
                        
                        if 'types' in df.columns:
                            df_export['type'] = df['types'].apply(
                                lambda x: x[0].get('name', 'Entity') if x and len(x) > 0 else 'Entity'
                            )
                        
                        df_export.columns = ['Entity Name', 'Match Score', 'ICIJ ID'] + (['Type'] if 'types' in df.columns else [])
                        df_export['Search Query'] = query
                        df_export['Search Date'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        df_export['ICIJ Link'] = df_export['ICIJ ID'].apply(
                            lambda x: f"{ICIJ_NODE_URL}{x}"
                        )
                        
                        # Export options
                        st.markdown("### Export Options")
                        col_csv, col_excel, col_pdf = st.columns(3)
                        
                        with col_csv:
                            st.download_button(
                                label="Download CSV",
                                data=df_export.to_csv(index=False),
                                file_name=f"icij_search_{query.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.csv",
                                mime="text/csv",
                                help="Download results as CSV file"
                            )
                        
                        with col_excel:
                            excel_data = create_excel_export(df_export, query, sources)
                            st.download_button(
                                label="Download Excel",
                                data=excel_data,
                                file_name=f"icij_search_{query.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                help="Download formatted Excel file"
                            )
                        
                        with col_pdf:
                            pdf_data = create_pdf_export(df_export[['Entity Name', 'Match Score', 'ICIJ ID']], query, sources)
                            st.download_button(
                                label="Download PDF",
                                data=pdf_data,
                                file_name=f"icij_search_{query.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf",
                                mime="application/pdf",
                                help="Download professional PDF report"
                            )
                        
                        with st.expander("View Data Table", expanded=False):
                            st.dataframe(
                                df_export,
                                use_container_width=True,
                                hide_index=True
                            )
                    
                    else:
                        st.warning("No results found matching the selected filters")
                        st.info("Try adjusting your filters or using a different search term.")
                
                else:
                    st.warning(f'No results found for "{query}"')
                    st.info("""
                    **Suggestions:**
                    - Check your spelling
                    - Use fewer words
                    - Search for company names or locations
                    - Try example searches from the sidebar
                    """)
        
        else:
            st.markdown("---")
            st.markdown("### Welcome")
            st.markdown("""
            Begin your search by entering a query above to explore the ICIJ Offshore Leaks Database.
            
            **Database Coverage:**
            - Panama Papers: 11.5 million leaked documents
            - Paradise Papers: 13.4 million documents  
            - Pandora Papers: 11.9 million documents
            - Bahamas Leaks: 1.3 million documents
            - Offshore Leaks: 2.5 million documents
            
            Use the sidebar for guidance and example searches.
            """)
    
    # ========================================================================
    # TAB 2: VISUALIZATIONS
    # ========================================================================
    with tab2:
        st.markdown("### Data Visualizations")
        
        history = get_search_history(50)
        
        if history:
            hist_df = pd.DataFrame(history, columns=['Query', 'Sources', 'Results', 'Date'])
            hist_df['Sources'] = hist_df['Sources'].apply(lambda x: x.split(',') if x else [])
            
            col_viz1, col_viz2 = st.columns(2)
            
            with col_viz1:
                st.markdown("#### Search Activity Over Time")
                hist_df['Date'] = pd.to_datetime(hist_df['Date'])
                daily_searches = hist_df.groupby(hist_df['Date'].dt.date).size().reset_index()
                daily_searches.columns = ['Date', 'Searches']
                
                fig_timeline = px.line(
                    daily_searches,
                    x='Date',
                    y='Searches',
                    title='Daily Search Activity',
                    markers=True
                )
                fig_timeline.update_traces(line_color=COLORS['primary_red'], marker=dict(size=8, color=COLORS['dark_red']))
                fig_timeline.update_layout(
                    plot_bgcolor='white',
                    paper_bgcolor='white',
                    font=dict(color=COLORS['dark_gray'], family='Gothic A1')
                )
                st.plotly_chart(fig_timeline, use_container_width=True)
            
            with col_viz2:
                st.markdown("#### Source Distribution")
                all_sources = []
                for sources_list in hist_df['Sources']:
                    if sources_list and sources_list != ['']:
                        all_sources.extend(sources_list)
                
                if all_sources:
                    source_counts = pd.Series(all_sources).value_counts().reset_index()
                    source_counts.columns = ['Source', 'Count']
                    
                    fig_sources = px.pie(
                        source_counts,
                        values='Count',
                        names='Source',
                        title='Most Searched Data Sources',
                        color_discrete_sequence=[COLORS['primary_red'], COLORS['dark_red'], COLORS['dark_gray'], COLORS['medium_gray'], COLORS['light_red']]
                    )
                    st.plotly_chart(fig_sources, use_container_width=True)
                else:
                    st.info("No source-specific searches yet")
            
            st.markdown("#### Results Distribution")
            fig_results = px.histogram(
                hist_df,
                x='Results',
                nbins=20,
                title='Distribution of Search Results',
                labels={'Results': 'Number of Results', 'count': 'Frequency'}
            )
            fig_results.update_traces(marker_color=COLORS['primary_red'])
            fig_results.update_layout(
                plot_bgcolor='white',
                paper_bgcolor='white',
                font=dict(color=COLORS['dark_gray'], family='Gothic A1')
            )
            st.plotly_chart(fig_results, use_container_width=True)
            
            st.markdown("#### Top Searches")
            top_queries = hist_df['Query'].value_counts().head(10).reset_index()
            top_queries.columns = ['Query', 'Search Count']
            
            fig_top = px.bar(
                top_queries,
                x='Search Count',
                y='Query',
                orientation='h',
                title='Most Frequent Queries',
                color='Search Count',
                color_continuous_scale=['#F5F5F5', COLORS['light_red'], COLORS['primary_red'], COLORS['dark_red']]
            )
            fig_top.update_layout(
                plot_bgcolor='white',
                paper_bgcolor='white',
                font=dict(color=COLORS['dark_gray'], family='Gothic A1'),
                showlegend=False
            )
            st.plotly_chart(fig_top, use_container_width=True)
        else:
            st.info("No search data available yet. Start searching to see visualizations of your activity.")
    
    # ========================================================================
    # TAB 3: COMPARE
    # ========================================================================
    with tab3:
        st.markdown("### Entity Comparison")
        
        if st.session_state.comparison_list:
            st.markdown(f"**Comparing {len(st.session_state.comparison_list)} entities**")
            
            if st.button("Clear All"):
                st.session_state.comparison_list = []
                st.rerun()
            
            comparison_data = []
            for entity in st.session_state.comparison_list:
                entity_type = entity.get('types', [{}])[0].get('name', 'Entity') if entity.get('types') else 'Entity'
                comparison_data.append({
                    'Name': entity.get('name', 'Unknown'),
                    'Type': entity_type,
                    'Match Score': entity.get('score', 0),
                    'ID': entity.get('id', ''),
                    'Description': entity.get('description', 'N/A')
                })
            
            comp_df = pd.DataFrame(comparison_data)
            
            st.markdown("#### Score Comparison")
            fig_compare = px.bar(
                comp_df,
                x='Name',
                y='Match Score',
                color='Type',
                title='Match Score Comparison',
                color_discrete_sequence=[COLORS['primary_red'], COLORS['dark_red'], COLORS['dark_gray'], COLORS['medium_gray']]
            )
            fig_compare.update_layout(
                plot_bgcolor='white',
                paper_bgcolor='white',
                font=dict(color=COLORS['dark_gray'], family='Gothic A1')
            )
            st.plotly_chart(fig_compare, use_container_width=True)
            
            st.markdown("#### Detailed Comparison Table")
            st.dataframe(comp_df, use_container_width=True, hide_index=True)
            
            st.markdown("#### Quick Links")
            for idx, entity in enumerate(st.session_state.comparison_list, 1):
                entity_id = entity.get('id', '')
                entity_name = entity.get('name', 'Unknown')
                col_link, col_remove = st.columns([5, 1])
                with col_link:
                    st.markdown(f"{idx}. [{entity_name}]({ICIJ_NODE_URL}{entity_id})")
                with col_remove:
                    if st.button("Remove", key=f"rem_{idx}"):
                        st.session_state.comparison_list.pop(idx-1)
                        st.rerun()
        else:
            st.info("Add entities from the Search tab to compare them here.")
            st.markdown("""
            **How to use:**
            1. Navigate to the Search tab
            2. Perform a search
            3. Click the 'Add' button next to results you want to compare
            4. Return to this tab to view the comparison
            """)
    
    # ========================================================================
    # TAB 4: SAVED & HISTORY
    # ========================================================================
    with tab4:
        st.markdown("### Search History & Saved Searches")
        
        col_hist, col_saved = st.columns(2)
        
        with col_hist:
            st.markdown("#### Recent Searches")
            history = get_search_history(15)
            if history:
                for h in history:
                    query_text, sources_str, count, date_str = h
                    sources_list = sources_str.split(',') if sources_str else []
                    sources_display = ', '.join(sources_list) if sources_list and sources_list != [''] else 'All sources'
                    
                    st.markdown(
                        f"""
                        <div class="comparison-card">
                            <strong>{query_text}</strong><br>
                            <small>Sources: {sources_display} | Results: {count} | Date: {date_str}</small>
                        </div>
                        """,
                        unsafe_allow_html=True
                    )
            else:
                st.info("No search history yet. Start searching to see your history here.")
        
        with col_saved:
            st.markdown("#### Saved Searches")
            
            with st.expander("Save New Search", expanded=False):
                save_name = st.text_input("Search Name", key="save_name")
                save_query = st.text_input("Query", key="save_query")
                save_sources = st.multiselect(
                    "Sources",
                    DATA_SOURCES,
                    key="save_sources"
                )
                save_notes = st.text_area("Notes (optional)", key="save_notes")
                
                if st.button("Save This Search"):
                    if save_name and save_query:
                        if save_search(save_name, save_query, save_sources, save_notes):
                            st.success(f"Saved '{save_name}'")
                            st.rerun()
                    else:
                        st.warning("Please provide both name and query")
            
            saved = get_saved_searches()
            if saved:
                for s in saved:
                    search_id, name, query_text, sources_str, notes, created_str = s
                    sources_list = sources_str.split(',') if sources_str else []
                    sources_display = ', '.join(sources_list) if sources_list and sources_list != [''] else 'All sources'
                    
                    col_display, col_action = st.columns([4, 1])
                    with col_display:
                        st.markdown(
                            f"""
                            <div class="comparison-card">
                                <strong>{name}</strong><br>
                                <small>Query: {query_text}</small><br>
                                <small>Sources: {sources_display} | Date: {created_str}</small>
                                {f'<br><small>Notes: {notes}</small>' if notes else ''}
                            </div>
                            """,
                            unsafe_allow_html=True
                        )
                    with col_action:
                        if st.button("Delete", key=f"del_{search_id}"):
                            if delete_saved_search(search_id):
                                st.success("Deleted")
                                st.rerun()
            else:
                st.info("No saved searches yet. Use the form above to save your favorite searches.")
    
    # Footer
    st.markdown("---")
    st.markdown(
        f"<p style='text-align: center; color: {COLORS['medium_gray']}; font-size: 0.9rem;'>Data provided by the International Consortium of Investigative Journalists (ICIJ)</p>",
        unsafe_allow_html=True
    )


if __name__ == "__main__":
    main()

